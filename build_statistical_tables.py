#!/usr/bin/env python3
"""Build the aggregate supplementary workbook from typed table JSON.

This public implementation uses openpyxl only. It contains no manuscript or
response-letter generation logic and never reads participant-level records.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


SHEET_NAMES = [
    "README", "S1_Cohort_Map", "S2_Continuous", "S3_Transitions",
    "S4_Threshold_Audit", "S5_State_Structure", "S6_Reliability",
    "S7_Interval_Lengths", "S8_IPCW", "S9_Mortality", "S10_Adjustment",
    "S11_Prediction", "S12_Absolute_Risks", "S13_Sensitivities",
    "S14_Model_Failures", "S15_Model_Completion", "S16_Additional_Analyses",
    "S17_Reference_Locked", "S18_Reconciliation_Reliability",
]


def safe_table_name(index: int, title: str) -> str:
    stem = re.sub(r"[^A-Za-z0-9_]", "_", title)[:150].strip("_")
    return f"Table_{index:02d}_{stem}"[:250]


def width_for(header: object) -> float:
    text = str(header).lower()
    if text in {"value", "note"}:
        return 48
    if any(key in text for key in ("reason", "covariate", "release", "cognition", "function", "source", "estimand", "cohorts")):
        return 34
    if any(key in text for key in ("outcome", "analysis", "term", "exposure", "definition", "status", "record")):
        return 24
    return 15


def build(input_path: Path, output_path: Path) -> None:
    payload = json.loads(input_path.read_text(encoding="utf-8"))
    sheets = payload.get("sheets", [])
    if len(sheets) != len(SHEET_NAMES):
        raise ValueError(f"Expected {len(SHEET_NAMES)} table specifications, found {len(sheets)}")

    workbook = Workbook()
    workbook.remove(workbook.active)
    for index, (name, spec) in enumerate(zip(SHEET_NAMES, sheets, strict=True), start=1):
        sheet = workbook.create_sheet(name)
        columns = list(spec["columns"])
        rows = list(spec["rows"])
        last_col = get_column_letter(len(columns))

        sheet.merge_cells(f"A1:{last_col}1")
        sheet["A1"] = spec["title"]
        sheet.merge_cells(f"A2:{last_col}2")
        sheet["A2"] = spec["note"]
        sheet.merge_cells(f"A3:{last_col}3")
        sheet["A3"] = f"Source: {spec['source']}"
        for col_index, value in enumerate(columns, start=1):
            sheet.cell(4, col_index, value)
        for row_index, row in enumerate(rows, start=5):
            for col_index, value in enumerate(row, start=1):
                sheet.cell(row_index, col_index, value)

        sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
        sheet["A1"].font = Font(color="FFFFFF", bold=True, size=14)
        sheet["A2"].fill = PatternFill("solid", fgColor="D9EAF7")
        sheet["A2"].font = Font(color="203864", size=10)
        sheet["A3"].fill = PatternFill("solid", fgColor="F2F2F2")
        sheet["A3"].font = Font(color="666666", italic=True, size=9)
        for cell in sheet[4]:
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.font = Font(color="FFFFFF", bold=True, size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        for row in sheet.iter_rows(min_row=5):
            for cell in row:
                cell.font = Font(color="222222", size=9)
                cell.alignment = Alignment(wrap_text=width_for(columns[cell.column - 1]) >= 24, vertical="top")
        for col_index, header in enumerate(columns, start=1):
            sheet.column_dimensions[get_column_letter(col_index)].width = width_for(header)
        sheet.row_dimensions[1].height = 24
        sheet.row_dimensions[2].height = 36
        sheet.row_dimensions[3].height = 22
        sheet.row_dimensions[4].height = 30
        sheet.freeze_panes = "B5" if len(columns) > 5 else "A5"
        sheet.sheet_view.showGridLines = False

        if rows:
            table = Table(displayName=safe_table_name(index, spec["title"]), ref=f"A4:{last_col}{4 + len(rows)}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False,
            )
            sheet.add_table(table)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"wrote {output_path} with {len(sheets)} sheets")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, default=Path(__file__).resolve().with_name("statistical_tables.json"))
    parser.add_argument("--output", type=Path, default=Path(__file__).resolve().with_name("Supplementary_Appendix_2_Tables_rebuilt.xlsx"))
    args = parser.parse_args()
    build(args.input, args.output)


if __name__ == "__main__":
    main()
